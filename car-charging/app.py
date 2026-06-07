"""Car Charging Costs — a Streamlit dashboard over EV charger meterdata.

Reads the charger's CSV export (upload one, use the bundled sample, or fetch
from the charger on your local network) and shows cost/energy KPIs, monthly
trends, charging habits, day/night tariff analysis, a per-car breakdown, the
session log, and Excel/CSV downloads.

Run with:  streamlit run app.py
"""

from __future__ import annotations

import io
import json
import os
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

from charging import (
    daily_summary,
    load_sessions,
    monthly_by_car,
    monthly_effective_price,
    monthly_summary,
    parse_metadata,
    split_peak_offpeak,
)

DEFAULT_CSV = Path(__file__).parent / "data" / "sample_meterdata.csv"
CHARGER_URL = "http://pblr-0012237.local/charging-history"

# A curated, harmonious qualitative palette so each car keeps one colour across
# every chart in the app (the actual map is built once cars are known).
CAR_PALETTE = [
    "#2a9d8f", "#e76f51", "#264653", "#e9c46a",
    "#8ab17d", "#5b8e7d", "#bc6c25", "#6d597a",
]
INK = "#1f2a44"
OFFPEAK_C, PEAK_C = "#2a9d8f", "#e9c46a"

st.set_page_config(page_title="Car Charging Costs", page_icon="🔌", layout="wide")

# Lightweight CSS: hero banner, KPI cards, tidier tabs/spacing.
st.markdown(
    """
    <style>
      .block-container {padding-top: 2.2rem;}
      .hero {background: linear-gradient(135deg,#264653 0%,#2a9d8f 100%);
             color:#fff; padding:18px 24px; border-radius:16px; margin-bottom:14px;}
      .hero h1 {margin:0; font-size:1.55rem; font-weight:700;}
      .hero p  {margin:.25rem 0 0; opacity:.85; font-size:.95rem;}
      div[data-testid="stMetric"] {background:#f7f9fb; border:1px solid #e6eaf0;
             border-radius:14px; padding:14px 16px;}
      div[data-testid="stMetricValue"] {font-size:1.35rem;}
      button[data-baseweb="tab"] {font-size:1rem; font-weight:600;}
    </style>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<div class="hero"><h1>🔌 Car Charging Costs</h1>'
    "<p>EV charging cost &amp; energy dashboard</p></div>",
    unsafe_allow_html=True,
)


def style_fig(fig, *, legend: bool = True):
    """Apply one consistent, professional look to every Plotly figure."""
    fig.update_layout(
        template="plotly_white",
        font=dict(family="Inter, Segoe UI, Helvetica, sans-serif", size=13, color=INK),
        margin=dict(l=10, r=10, t=30, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=""),
        hoverlabel=dict(font_size=12),
        showlegend=legend,
        bargap=0.18,
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(showgrid=True, gridcolor="#eef1f6", zeroline=False)
    return fig


def _number_format(col: str) -> str | None:
    name = str(col).lower()
    if "per_kwh" in name:
        return "€#,##0.000"
    if "cost" in name or "€" in str(col) or "eur" in name:
        return "€#,##0.00"
    if "kwh" in name:
        return "#,##0.0"
    if "km" in name:
        return "#,##0"
    return None


def to_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Write a formatted, multi-sheet workbook with native charts."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        for name, frame in sheets.items():
            frame.to_excel(xl, sheet_name=name, index=False)
        wb = xl.book
        header_fill = PatternFill("solid", fgColor="1F2A44")
        header_font = Font(bold=True, color="FFFFFF")
        for name, frame in sheets.items():
            ws = wb[name]
            for col_idx, col in enumerate(frame.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill, cell.font = header_fill, header_font
                cell.alignment = Alignment(horizontal="center")
                values = [str(v) for v in frame[col]] if len(frame) else []
                width = max([len(str(col))] + [len(v) for v in values]) + 2
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 11), 42)
                fmt = _number_format(col)
                if fmt and len(frame):
                    for row in range(2, len(frame) + 2):
                        ws.cell(row=row, column=col_idx).number_format = fmt
            ws.freeze_panes = "A2"
            if len(frame):
                ws.auto_filter.ref = ws.dimensions

        def add_bar(sheet: str, value_col: int, title: str, anchor: str) -> None:
            ws = wb[sheet]
            if ws.max_row < 2:
                return
            chart = BarChart()
            chart.type, chart.title, chart.legend = "col", title, None
            chart.height, chart.width = 8, 18
            chart.add_data(
                Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row),
                titles_from_data=True,
            )
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            ws.add_chart(chart, anchor)

        add_bar("Monthly", 4, "Cost per month (€)", "F2")
        add_bar("By car", 4, "Cost by car (€)", "F2")
        add_bar("Cost per 100km", 5, "€ per 100 km", "G2")
    return buf.getvalue()


def p1_details_df(reading) -> pd.DataFrame:
    """Every field the P1 meter returned, as a tidy two-column table."""
    items = list((reading.details or {}).items())
    return pd.DataFrame(items, columns=["Field", "Value"])


def home_vs_car_df(reading, home_price: float, car_kwh: float, car_cost: float) -> pd.DataFrame:
    """A small home-energy vs car-charging comparison table."""
    rows: list[list] = []
    if reading.import_kwh is not None:
        rows.append(["Home imported (lifetime, kWh)", round(reading.import_kwh, 1)])
        rows.append(["Home cost (lifetime, €)", round(reading.import_kwh * home_price, 2)])
    if reading.export_kwh is not None:
        rows.append(["Home exported (lifetime, kWh)", round(reading.export_kwh, 1)])
    rows.append(["Car charging (selected period, kWh)", round(car_kwh, 1)])
    rows.append(["Car cost (selected period, €)", round(car_cost, 2)])
    if reading.import_kwh:
        rows.append(["Car as % of home lifetime import", round(car_kwh / reading.import_kwh * 100, 1)])
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _setting(name: str, default: str = "") -> str:
    """Read a setting from Streamlit secrets, falling back to env vars."""
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return os.environ.get(name, default)


SECRETS_PATH = Path(__file__).parent / ".streamlit" / "secrets.toml"


def _save_secrets(updates: dict) -> None:
    """Persist key = "value" settings to the local (git-ignored) secrets.toml."""
    SECRETS_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = (
        SECRETS_PATH.read_text(encoding="utf-8").splitlines()
        if SECRETS_PATH.exists() else []
    )
    for key, val in updates.items():
        escaped = str(val).replace("\\", "\\\\").replace('"', '\\"')
        newline = f'{key} = "{escaped}"'
        for i, ln in enumerate(lines):
            stripped = ln.lstrip()
            if not stripped.startswith("#") and (
                stripped.startswith(f"{key} ") or stripped.startswith(f"{key}=")
            ):
                lines[i] = newline
                break
        else:
            lines.append(newline)
    SECRETS_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _forget_secrets(keys: list) -> None:
    """Remove the given keys from the local secrets.toml, keeping comments."""
    if not SECRETS_PATH.exists():
        return
    kept = []
    for ln in SECRETS_PATH.read_text(encoding="utf-8").splitlines():
        stripped = ln.lstrip()
        if not stripped.startswith("#") and any(
            stripped.startswith(f"{k} ") or stripped.startswith(f"{k}=") for k in keys
        ):
            continue
        kept.append(ln)
    SECRETS_PATH.write_text("\n".join(kept) + "\n", encoding="utf-8")


def _remember_controls(scope: str, values: dict) -> None:
    """Render 'Remember on this PC' / 'Forget' buttons for local persistence."""
    st.caption(
        "💾 stores these on this computer in `.streamlit/secrets.toml` "
        "(git-ignored, never uploaded), so they're filled in next time."
    )
    cols = st.columns(2)
    if cols[0].button("💾 Remember on this PC", key=f"save_{scope}"):
        try:
            _save_secrets({k: v for k, v in values.items() if v})
            st.success("Saved on this computer — it'll be here next time.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Couldn't save here (read-only host?): {exc}")
    if cols[1].button("🗑️ Forget", key=f"forget_{scope}"):
        try:
            _forget_secrets(list(values.keys()))
            st.info("Removed from this computer.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Couldn't update secrets: {exc}")


REPO_ZIP_URL = "https://github.com/gerrit0492-create/Obsidian/archive/refs/heads/main.zip"


def _self_update() -> str:
    """Download the latest car-charging files from GitHub over this folder.

    Local secrets are preserved; restart the app afterwards to load the changes.
    """
    import io as _io
    import shutil
    import urllib.request
    import zipfile

    here = Path(__file__).parent
    with urllib.request.urlopen(REPO_ZIP_URL, timeout=60) as resp:
        data = resp.read()
    zf = zipfile.ZipFile(_io.BytesIO(data))
    prefix = "Obsidian-main/car-charging/"
    count = 0
    for name in zf.namelist():
        if not name.startswith(prefix) or name.endswith("/"):
            continue
        rel = name[len(prefix):]
        if rel.startswith(".streamlit/secrets.toml"):  # never clobber local secrets
            continue
        dest = here / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        with zf.open(name) as src, open(dest, "wb") as out:
            shutil.copyfileobj(src, out)
        count += 1
    return f"Updated {count} files from GitHub."


def _require_password() -> None:
    expected = _setting("APP_PASSWORD")
    if not expected or st.session_state.get("authed"):
        return
    pw = st.text_input("Password", type="password")
    if pw and pw == expected:
        st.session_state["authed"] = True
        return
    if pw:
        st.error("Wrong password.")
    st.stop()


# Card -> car mapping used unless overridden by the CAR_MAP setting. Keyed by the
# friendly card label (or RFID id); see charging.load_sessions.
_DEFAULT_CAR_MAP = {"ANWB": "Citroën ëC3", "ANWB Gerrit": "Kia Sorento"}


def _car_map() -> dict:
    raw = _setting("CAR_MAP")
    if not raw:
        return dict(_DEFAULT_CAR_MAP)
    try:
        return {str(k): str(v) for k, v in json.loads(raw).items()}
    except Exception:
        return dict(_DEFAULT_CAR_MAP)


def _car_wltp() -> dict:
    """Map car name -> manufacturer WLTP consumption (kWh/100km), from CAR_WLTP."""
    raw = _setting("CAR_WLTP")
    if not raw:
        return {}
    try:
        return {str(k): float(v) for k, v in json.loads(raw).items()}
    except Exception:
        return {}


_require_password()

# --- App update ------------------------------------------------------------
with st.sidebar.expander("⚙️ App"):
    st.caption("Get the latest version from GitHub. Your saved settings are kept.")
    if st.button("⬆️ Update app"):
        try:
            msg = _self_update()
            st.success(f"{msg} Close this window and start the app again to apply.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Update failed: {exc}")

# --- How to start / auto-start ---------------------------------------------
with st.sidebar.expander("▶️ Start / auto-start"):
    st.markdown(
        "**Start by hand**\n"
        "- Double-click **`start.bat`** in the `car-charging` folder, or\n"
        "- open `cmd` in that folder and run `python -m streamlit run app.py`.\n"
        "- The dashboard opens at http://localhost:8501. Close the window to stop it.\n\n"
        "**Start automatically at login** (Windows)\n"
        "- Run **`install-autostart.bat`** once — it launches the dashboard every "
        "time you log in.\n"
        "- Undo it any time with **`uninstall-autostart.bat`**.\n\n"
        "Live P1 + charger panels need this laptop on the **same Wi-Fi** as the devices."
    )

# --- Data source -----------------------------------------------------------
st.sidebar.header("Data")
uploaded = st.sidebar.file_uploader("Upload meterdata CSV", type=["csv"])

source_text: str | None = None
if uploaded is not None:
    source_text = uploaded.getvalue().decode("utf-8", "replace")
else:
    with st.sidebar.expander("Fetch from charger (same network only)"):
        url = st.text_input("Charger URL", value=CHARGER_URL)
        if st.button("Fetch now"):
            try:
                import requests

                resp = requests.get(url, timeout=10)
                resp.raise_for_status()
                source_text = resp.text
                st.success("Fetched from charger.")
            except Exception as exc:  # noqa: BLE001
                st.error(
                    f"Couldn't reach the charger: {exc}. This only works when the app "
                    "runs on the same network as the charger."
                )
    if source_text is None and DEFAULT_CSV.exists():
        source_text = DEFAULT_CSV.read_text(encoding="utf-8")
        st.sidebar.caption("Showing bundled sample data — upload a CSV to use your own.")

if not source_text:
    st.info("Upload a meterdata CSV in the sidebar to begin.")
    st.stop()

meta = parse_metadata(source_text)
df = load_sessions(io.StringIO(source_text), car_map=_car_map())

caption = []
if meta.serial:
    caption.append(f"Charger {meta.serial}")
if meta.period_from:
    caption.append(f"{meta.period_from[:10]} → {meta.period_to[:10]}")
if caption:
    st.caption(" · ".join(caption))

# --- Filters ---------------------------------------------------------------
st.sidebar.header("Filters")
min_d, max_d = df["start"].min().date(), df["start"].max().date()
date_range = st.sidebar.date_input(
    "Date range", (min_d, max_d), min_value=min_d, max_value=max_d
)
cars = sorted(df["car"].dropna().unique())
chosen = st.sidebar.multiselect("Cars", cars, default=cars)

# One colour per car, stable across every chart and regardless of the filter.
CAR_COLOR = {car: CAR_PALETTE[i % len(CAR_PALETTE)] for i, car in enumerate(cars)}

# Day/night (dal) tariff settings — drive the "Best time to charge" view.
st.sidebar.header("Tariff (day/night)")
st.sidebar.caption(
    "Your real home prices. Both default to the charger's logged price; set a "
    "lower dal price to estimate day/night savings."
)
# Default to the charger's own price so figures match the CSV until you set a split.
default_price = round(float(meta.price_per_kwh), 2) if pd.notna(meta.price_per_kwh) else 0.30
dal_start = st.sidebar.number_input("Off-peak starts (hour)", 0, 23, 23)
dal_end = st.sidebar.number_input("Off-peak ends (hour)", 0, 23, 7)
weekend_off = st.sidebar.checkbox("Weekends fully off-peak", value=True)
price_normaal = st.sidebar.number_input(
    "Peak €/kWh (normaal)", min_value=0.0, value=default_price, step=0.01, format="%.2f"
)
price_dal = st.sidebar.number_input(
    "Off-peak €/kWh (dal)", min_value=0.0, value=default_price, step=0.01, format="%.2f"
)


def hour_is_off(hour: int) -> bool:
    """Off-peak by clock hour only (ignores the weekend rule), for shading charts."""
    if dal_start <= dal_end:
        return dal_start <= hour < dal_end
    return hour >= dal_start or hour < dal_end


mask = df["car"].isin(chosen)
if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
    start_d, end_d = date_range
    mask &= (df["date"] >= start_d) & (df["date"] <= end_d)
fdf = df[mask].copy()

if fdf.empty:
    st.warning("No sessions match the current filters.")
    st.stop()

# --- KPIs (always visible above the tabs) ----------------------------------
total_cost = fdf["cost"].sum()
total_kwh = fdf["energy_kwh"].sum()
n_sessions = int((fdf["energy_kwh"] > 0).sum())
avg_cost = total_cost / n_sessions if n_sessions else 0.0
eur_per_kwh = total_cost / total_kwh if total_kwh else 0.0
total_hours = fdf["duration_h"].sum()

# Month-over-month deltas (latest full month vs the previous one), shown on the cards.
_msum = monthly_summary(fdf)
d_sessions = d_energy = d_cost = None
if len(_msum) >= 2:
    last, prev = _msum.iloc[-1], _msum.iloc[-2]
    d_sessions = f"{int(last['sessions'] - prev['sessions']):+d}"
    d_energy = f"{last['energy_kwh'] - prev['energy_kwh']:+,.1f} kWh"
    d_cost = f"€{last['cost'] - prev['cost']:+,.2f}"

k = st.columns(5)
k[0].metric("Charging sessions", f"{n_sessions}", delta=d_sessions)
k[1].metric("Energy", f"{total_kwh:,.1f} kWh", delta=d_energy)
k[2].metric("Cost", f"€{total_cost:,.2f}", delta=d_cost, delta_color="inverse")
k[3].metric("Avg €/session", f"€{avg_cost:,.2f}")
k[4].metric("Effective €/kWh", f"€{eur_per_kwh:,.3f}")
if d_cost is not None:
    st.caption("Δ on cards = latest month vs the previous month.")

tab_overview, tab_home, tab_habits, tab_charging, tab_cars, tab_data = st.tabs(
    ["📊 Overview", "🏠 Home", "🕒 Habits", "⚡ Charging", "🚗 Cars", "📑 Data"]
)

# ===========================================================================
# Overview
# ===========================================================================
with tab_overview:
    msum = monthly_summary(fdf)
    mbc = monthly_by_car(fdf)
    left, right = st.columns(2)
    with left:
        st.subheader("Cost per month")
        fig = px.bar(
            mbc, x="month", y="cost", color="car", text_auto=".2f",
            color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_traces(hovertemplate="%{x}<br>%{fullData.name}: €%{y:.2f}<extra></extra>")
        fig.update_layout(barmode="stack", yaxis_title="€", xaxis_title="", yaxis_tickprefix="€")
        st.plotly_chart(style_fig(fig), use_container_width=True)
    with right:
        st.subheader("Energy per month")
        fig = px.bar(
            mbc, x="month", y="energy_kwh", text_auto=".0f", color="car",
            color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_traces(hovertemplate="%{x}<br>%{fullData.name}: %{y:.1f} kWh<extra></extra>")
        fig.update_layout(barmode="stack", yaxis_title="kWh", xaxis_title="", yaxis_ticksuffix=" kWh")
        st.plotly_chart(style_fig(fig), use_container_width=True)

    left, right = st.columns([3, 2])
    with left:
        st.subheader("Cumulative cost by car")
        cum = fdf.sort_values("start").copy()
        cum["cumulative_cost"] = cum.groupby("car")["cost"].cumsum()
        fig = px.area(
            cum, x="start", y="cumulative_cost", color="car",
            color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_traces(hovertemplate="%{x|%d %b %Y}<br>%{fullData.name}: €%{y:.2f}<extra></extra>")
        fig.update_layout(yaxis_title="€", xaxis_title="", yaxis_tickprefix="€")
        st.plotly_chart(style_fig(fig), use_container_width=True)
    with right:
        st.subheader("Cost share by car")
        by_car = (
            fdf.groupby("car")
            .agg(sessions=("session", "count"), energy_kwh=("energy_kwh", "sum"), cost=("cost", "sum"))
            .reset_index()
        )
        fig = px.pie(
            by_car, names="car", values="cost", hole=0.55,
            color="car", color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_traces(
            textposition="inside",
            texttemplate="%{label}<br>€%{value:.0f}",
            hovertemplate="%{label}: €%{value:.2f} (%{percent})<extra></extra>",
        )
        fig.update_layout(
            annotations=[dict(text=f"€{total_cost:,.0f}", x=0.5, y=0.5, font_size=18, showarrow=False)]
        )
        st.plotly_chart(style_fig(fig, legend=False), use_container_width=True)

# ===========================================================================
# Habits + best time to charge
# ===========================================================================
with tab_habits:
    st.subheader("When you charge")
    hab = fdf.dropna(subset=["start"]).copy()
    hab["hour"] = hab["start"].dt.hour
    hab["weekday"] = hab["start"].dt.day_name()
    weekday_order = [
        "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"
    ]
    left, right = st.columns([3, 1])
    with left:
        fig = px.density_heatmap(
            hab, x="hour", y="weekday", z="energy_kwh", histfunc="sum",
            nbinsx=24, category_orders={"weekday": weekday_order},
            color_continuous_scale="Tealgrn",
        )
        fig.update_traces(hovertemplate="%{y} at %{x}:00<br>%{z:.1f} kWh<extra></extra>")
        fig.update_layout(
            xaxis_title="Hour of day", yaxis_title="",
            coloraxis_colorbar=dict(title="kWh", thickness=12),
        )
        fig.update_xaxes(dtick=2, range=[-0.5, 23.5])
        # Outline the off-peak (dal) hours so habits can be read against the tariff.
        for h in [h for h in range(24) if hour_is_off(h)]:
            fig.add_vrect(x0=h - 0.5, x1=h + 0.5, fillcolor=OFFPEAK_C, opacity=0.10, line_width=0)
        st.plotly_chart(style_fig(fig, legend=False), use_container_width=True)
    with right:
        total_hab = hab["energy_kwh"].sum()
        start_frac = hab["start"].dt.hour + hab["start"].dt.minute / 60.0
        if not start_frac.empty:
            avg_h = start_frac.mean()
            st.metric("Typical plug-in time", f"{int(avg_h):02d}:{int((avg_h % 1) * 60):02d}")
        night = hab[(hab["hour"] >= 23) | (hab["hour"] < 7)]["energy_kwh"].sum()
        st.metric("Charged 23:00–07:00", f"{night / total_hab * 100 if total_hab else 0:.0f}%")
        weekend = hab[hab["weekday"].isin(["Saturday", "Sunday"])]["energy_kwh"].sum()
        st.metric("Weekend share", f"{weekend / total_hab * 100 if total_hab else 0:.0f}%")

    st.divider()
    st.subheader("⏰ Best time to charge")
    split = split_peak_offpeak(
        fdf, dal_start=int(dal_start), dal_end=int(dal_end), weekend_offpeak=weekend_off
    )
    off_kwh = split["energy_offpeak"].sum()
    peak_kwh = split["energy_peak"].sum()
    tou_total = off_kwh + peak_kwh
    off_share = off_kwh / tou_total * 100 if tou_total else 0.0
    home_cost = off_kwh * price_dal + peak_kwh * price_normaal
    all_off_cost = tou_total * price_dal
    savings = peak_kwh * (price_normaal - price_dal)
    window_txt = f"{int(dal_start):02d}:00–{int(dal_end):02d}:00" + (
        " + weekends" if weekend_off else ""
    )

    c = st.columns(4)
    c[0].metric("Charged off-peak", f"{off_share:.0f}%")
    c[1].metric("Cost at home tariff", f"€{home_cost:,.2f}")
    c[2].metric("If all off-peak", f"€{all_off_cost:,.2f}")
    c[3].metric("Potential saving", f"€{savings:,.2f}")
    st.caption(
        "What-if on your **home** day/night tariff (sidebar) — separate from the flat "
        "price the charger logs in the CSV."
    )

    if abs(price_normaal - price_dal) < 1e-9:
        st.info(
            "Set a lower **off-peak (dal)** price than **peak (normaal)** in the sidebar "
            "to estimate day/night savings — both currently match the charger's flat price, "
            "so there's nothing to compare yet."
        )
    elif savings > 0.01:
        st.success(
            f"**Charge during the dal window ({window_txt}).** You charged {off_share:.0f}% "
            f"off-peak — shifting the rest would save about €{savings:,.2f} at "
            f"€{price_normaal:.2f}/€{price_dal:.2f} per kWh (peak/off-peak)."
        )
    else:
        st.success(
            f"Nicely timed — almost all of your charging already falls in the dal window "
            f"({window_txt})."
        )

    left, right = st.columns(2)
    with left:
        st.caption("Off-peak vs peak energy by car")
        tou = (
            split.groupby("car")[["energy_offpeak", "energy_peak"]]
            .sum()
            .reset_index()
            .melt(id_vars="car", var_name="period", value_name="kwh")
        )
        tou["period"] = tou["period"].map({"energy_offpeak": "Off-peak", "energy_peak": "Peak"})
        fig = px.bar(
            tou, x="car", y="kwh", color="period", barmode="stack",
            color_discrete_map={"Off-peak": OFFPEAK_C, "Peak": PEAK_C},
        )
        fig.update_layout(yaxis_title="kWh", xaxis_title="", yaxis_ticksuffix=" kWh")
        st.plotly_chart(style_fig(fig), use_container_width=True)
    with right:
        st.caption("Energy by hour of day (green = off-peak)")
        hh = split.dropna(subset=["start"]).copy()
        hh["hour"] = hh["start"].dt.hour
        by_hour = (
            hh.groupby("hour")["energy_kwh"].sum().reindex(range(24), fill_value=0).reset_index()
        )
        by_hour["period"] = by_hour["hour"].map(lambda h: "Off-peak" if hour_is_off(h) else "Peak")
        fig = px.bar(
            by_hour, x="hour", y="energy_kwh", color="period",
            color_discrete_map={"Off-peak": OFFPEAK_C, "Peak": PEAK_C},
        )
        fig.update_layout(yaxis_title="kWh", xaxis_title="Hour of day", yaxis_ticksuffix=" kWh")
        fig.update_xaxes(dtick=2)
        st.plotly_chart(style_fig(fig), use_container_width=True)

    tou_summary = pd.DataFrame(
        {
            "Period": ["Off-peak (dal)", "Peak (normaal)", "Total"],
            "kWh": [round(off_kwh, 1), round(peak_kwh, 1), round(tou_total, 1)],
            "Price (€/kWh)": [price_dal, price_normaal, None],
            "Cost (€)": [
                round(off_kwh * price_dal, 2),
                round(peak_kwh * price_normaal, 2),
                round(home_cost, 2),
            ],
        }
    )

# ===========================================================================
# Charging behaviour
# ===========================================================================
with tab_charging:
    st.subheader("Charge speed")
    sp = fdf.dropna(subset=["power_kw"])
    left, right = st.columns(2)
    with left:
        st.caption("Power per session, spread by car")
        fig = px.box(
            sp, x="car", y="power_kw", color="car", points="all",
            color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_layout(yaxis_title="kW", xaxis_title="", yaxis_ticksuffix=" kW")
        st.plotly_chart(style_fig(fig, legend=False), use_container_width=True)
    with right:
        st.caption("Average power by car")
        by_car_power = sp.groupby("car")["power_kw"].mean().reset_index()
        fig = px.bar(
            by_car_power, x="car", y="power_kw", color="car", text_auto=".1f",
            color_discrete_map=CAR_COLOR, category_orders={"car": cars},
        )
        fig.update_traces(hovertemplate="%{x}: %{y:.2f} kW<extra></extra>")
        fig.update_layout(yaxis_title="kW", xaxis_title="", yaxis_ticksuffix=" kW")
        st.plotly_chart(style_fig(fig, legend=False), use_container_width=True)

    st.divider()
    st.subheader("Daily trend & forecast")
    daily = daily_summary(fdf)
    daily["cost_7d_avg"] = daily["cost"].rolling(7, min_periods=1).mean()
    daily_by_car = fdf.groupby(["date", "car"])["cost"].sum().reset_index()
    fig = px.bar(
        daily_by_car, x="date", y="cost", color="car",
        color_discrete_map=CAR_COLOR, category_orders={"car": cars},
    )
    fig.add_scatter(
        x=daily["date"], y=daily["cost_7d_avg"], mode="lines", name="7-day avg",
        line=dict(color=INK, width=2, dash="dot"),
    )
    fig.update_layout(barmode="stack", yaxis_title="€", xaxis_title="", yaxis_tickprefix="€")
    st.plotly_chart(style_fig(fig), use_container_width=True)

    span_days = (fdf["date"].max() - fdf["date"].min()).days + 1
    avg_daily_cost = total_cost / span_days if span_days else 0.0
    avg_daily_kwh = total_kwh / span_days if span_days else 0.0
    f = st.columns(3)
    f[0].metric("Avg cost / day", f"€{avg_daily_cost:,.2f}")
    f[1].metric("Projected / month", f"€{avg_daily_cost * 30:,.2f}")
    f[2].metric("Projected / year", f"€{avg_daily_cost * 365:,.0f}")
    st.caption(
        f"Run-rate projection from {avg_daily_kwh:,.1f} kWh/day over the selected "
        f"{span_days}-day period — not a seasonal forecast."
    )

    st.divider()
    st.subheader("Effective price per kWh over time")
    mprice = monthly_effective_price(fdf)
    mprice_car = (
        fdf.groupby(["month", "car"])
        .agg(energy_kwh=("energy_kwh", "sum"), cost=("cost", "sum"))
        .reset_index()
    )
    mprice_car["eur_per_kwh"] = mprice_car["cost"] / mprice_car["energy_kwh"]
    fig = px.line(
        mprice_car, x="month", y="eur_per_kwh", color="car", markers=True,
        color_discrete_map=CAR_COLOR, category_orders={"car": cars},
    )
    fig.update_traces(hovertemplate="%{x}<br>%{fullData.name}: €%{y:.3f}/kWh<extra></extra>")
    fig.update_layout(yaxis_title="€/kWh", xaxis_title="", yaxis_tickprefix="€")
    if pd.notna(meta.price_per_kwh):
        fig.add_hline(
            y=meta.price_per_kwh, line_dash="dot", line_color="#888",
            annotation_text=f"charger setting €{meta.price_per_kwh:.3f}",
            annotation_position="top left",
        )
    st.plotly_chart(style_fig(fig), use_container_width=True)

# ===========================================================================
# Per-car cost
# ===========================================================================
with tab_cars:
    st.subheader("Cost per 100 km (estimate)")
    st.caption(
        "Defaults to each car's manufacturer WLTP consumption (set via the CAR_WLTP "
        "secret); adjust for your real-world driving."
    )
    wltp = _car_wltp()
    eff_cols = st.columns(len(chosen) or 1)
    effs = {
        car: eff_cols[i].number_input(
            f"{car} — kWh/100km" + (" (WLTP)" if car in wltp else ""),
            min_value=1.0,
            value=float(wltp.get(car, 18.0)),
            step=0.5,
            key=f"eff_{car}",
        )
        for i, car in enumerate(chosen)
    }
    per100_rows = []
    for car in chosen:
        sub = fdf[fdf["car"] == car]
        energy = sub["energy_kwh"].sum()
        cost = sub["cost"].sum()
        eff = effs[car]
        km = energy / eff * 100 if eff else 0.0
        per100_rows.append(
            {
                "Car": car,
                "kWh": round(energy, 1),
                "Est. km": round(km),
                "Cost (€)": round(cost, 2),
                "€/100km": round(cost / km * 100, 2) if km else 0.0,
            }
        )
    per100 = pd.DataFrame(per100_rows)
    left, right = st.columns([2, 3])
    with left:
        st.dataframe(per100, use_container_width=True, hide_index=True)
    with right:
        if not per100.empty:
            fig = px.bar(
                per100, x="Car", y="€/100km", color="Car", text_auto=".2f",
                color_discrete_map=CAR_COLOR, category_orders={"Car": cars},
            )
            fig.update_traces(hovertemplate="%{x}: €%{y:.2f}/100km<extra></extra>")
            fig.update_layout(yaxis_title="€/100km", xaxis_title="", yaxis_tickprefix="€")
            st.plotly_chart(style_fig(fig, legend=False), use_container_width=True)

# ===========================================================================
# Home energy (HomeWizard P1)
# ===========================================================================
with tab_home:
    st.subheader("🏠 Home energy (HomeWizard P1)")
    st.caption(
        "Direct read works on your home network. To read it from anywhere (e.g. this "
        "hosted app), point the relay URL at a Home Assistant endpoint that republishes "
        "the meter JSON — see the README."
    )
    h1, h2, h3 = st.columns(3)
    hw_host = h1.text_input("P1 address (IP or hostname)", value=_setting("P1_HOST", "192.168.1.31"))
    hw_token = h2.text_input("Token (optional)", type="password", value=_setting("P1_TOKEN"))
    home_price = h3.number_input("Home €/kWh", min_value=0.0, value=0.35, step=0.01, format="%.2f")
    hw_remote = st.text_input(
        "…or remote relay URL (works away from home)",
        value=_setting("P1_REMOTE_URL"),
        placeholder="https://your-relay.example/p1.json",
    )
    _remember_controls("p1", {"P1_HOST": hw_host, "P1_TOKEN": hw_token, "P1_REMOTE_URL": hw_remote})

    if st.button("Read P1 now"):
        if not hw_host.strip() and not hw_remote.strip():
            st.error("Enter your P1 meter's address, or a remote relay URL.")
        else:
            try:
                from homewizard import fetch, fetch_url

                token = hw_token.strip() or None
                if hw_remote.strip():
                    st.session_state["p1_reading"] = fetch_url(hw_remote.strip(), token=token)
                else:
                    st.session_state["p1_reading"] = fetch(hw_host.strip(), token=token)
            except Exception as exc:  # noqa: BLE001
                st.error(
                    f"Couldn't read the P1 meter: {exc}. A direct address only works on the same "
                    "network (enable the Local API in the HomeWizard app); from elsewhere, use a "
                    "remote relay URL."
                )

    r = st.session_state.get("p1_reading")
    if r is not None:
        m = st.columns(4)
        m[0].metric("Live power", f"{r.active_power_w:,.0f} W" if r.active_power_w is not None else "—")
        m[1].metric("Imported (lifetime)", f"{r.import_kwh:,.0f} kWh" if r.import_kwh is not None else "—")
        m[2].metric("Exported (lifetime)", f"{r.export_kwh:,.0f} kWh" if r.export_kwh is not None else "—")
        if r.import_kwh is not None:
            m[3].metric("Home cost (lifetime)", f"€{r.import_kwh * home_price:,.0f}")

        if r.import_kwh:
            share = total_kwh / r.import_kwh * 100
            st.info(
                f"Car charging in view ({total_kwh:,.0f} kWh) is ~{share:.1f}% of the home's "
                f"lifetime imported energy ({r.import_kwh:,.0f} kWh). The home figure is lifetime, "
                "the car figure is the selected period — log P1 readings over time for a true "
                "side-by-side."
            )

        st.markdown("**Home vs car**")
        st.dataframe(
            home_vs_car_df(r, home_price, total_kwh, total_cost),
            use_container_width=True, hide_index=True,
        )

        det = p1_details_df(r)
        st.markdown(f"**All P1 data** — {len(det)} fields")
        st.dataframe(det, use_container_width=True, hide_index=True)
        with st.expander("Raw P1 JSON"):
            st.json(r.raw or {})

    st.divider()
    st.subheader("🔌 Live charger (Peblar)")
    st.caption(
        "Live state from the charger's local REST API — works on the home network. "
        "Enable the Local REST API in the charger's Advanced settings and paste the token. "
        "This is live state only; the session history comes from the CSV export."
    )
    c1, c2 = st.columns(2)
    chg_host = c1.text_input(
        "Charger address (IP or hostname)", value=_setting("CHARGER_HOST", "pblr-0012237.local")
    )
    chg_token = c2.text_input("API token", type="password", value=_setting("CHARGER_TOKEN"))
    _remember_controls("charger", {"CHARGER_HOST": chg_host, "CHARGER_TOKEN": chg_token})

    if st.button("Read charger now"):
        if not chg_host.strip() or not chg_token.strip():
            st.error("Enter the charger's address and its Local REST API token.")
        else:
            try:
                from peblar import fetch as fetch_charger

                r = fetch_charger(chg_host.strip(), chg_token.strip())
                m = st.columns(4)
                m[0].metric("Live power", f"{r.power_w:,.0f} W" if r.power_w is not None else "—")
                m[1].metric("This session", f"{r.session_kwh:,.2f} kWh" if r.session_kwh is not None else "—")
                m[2].metric("Lifetime", f"{r.total_kwh:,.0f} kWh" if r.total_kwh is not None else "—")
                m[3].metric("Status", r.cp_state or "—")
                if r.session_kwh and eur_per_kwh:
                    st.caption(
                        f"Current session ≈ €{r.session_kwh * eur_per_kwh:,.2f} at the effective "
                        f"€{eur_per_kwh:,.3f}/kWh from the sessions."
                    )
            except Exception as exc:  # noqa: BLE001
                st.error(
                    f"Couldn't read the charger: {exc}. This only works on the same network as "
                    "the charger, with the Local REST API enabled and a valid token."
                )


# ===========================================================================
# Data: session log, downloads
# ===========================================================================
with tab_data:
    st.subheader("Sessions")
    show = (
        fdf[["session", "start", "stop", "energy_kwh", "cost", "duration_h", "car", "validation"]]
        .sort_values("start", ascending=False)
        .rename(
            columns={
                "session": "Session",
                "start": "Start",
                "stop": "Stop",
                "energy_kwh": "kWh",
                "cost": "Cost (€)",
                "duration_h": "Hours",
                "car": "Car",
                "validation": "Validation",
            }
        )
    )
    st.dataframe(show, use_container_width=True, hide_index=True)

    summary = pd.DataFrame(
        {
            "Metric": [
                "Charging sessions", "Energy (kWh)", "Cost (€)", "Avg €/session",
                "Effective €/kWh", "Total charging hours", "Off-peak share (%)",
                "Period", "Cars",
            ],
            "Value": [
                n_sessions, round(total_kwh, 1), round(total_cost, 2), round(avg_cost, 2),
                round(eur_per_kwh, 3), round(total_hours, 1), round(off_share, 0),
                f"{fdf['date'].min()} → {fdf['date'].max()}", ", ".join(chosen),
            ],
        }
    )

    # Assemble a per-car context for the PDF report.
    car_share = split.groupby("car")[["energy_offpeak", "energy_peak"]].sum()
    per100_idx = per100.set_index("Car") if not per100.empty else None
    per_car_ctx = []
    for car in chosen:
        sub = fdf[fdf["car"] == car]
        e, cst = sub["energy_kwh"].sum(), sub["cost"].sum()
        ns = int((sub["energy_kwh"] > 0).sum())
        avgp = sub["power_kw"].mean()
        off = float(car_share.loc[car, "energy_offpeak"]) if car in car_share.index else 0.0
        pk = float(car_share.loc[car, "energy_peak"]) if car in car_share.index else 0.0
        metrics = [
            ("Sessions", f"{ns}"),
            ("Energy", f"{e:,.1f} kWh"),
            ("Cost", f"€{cst:,.2f}"),
            ("Avg power", f"{avgp:,.1f} kW" if pd.notna(avgp) else "—"),
            ("Effective €/kWh", f"€{cst / e:,.3f}" if e else "—"),
            ("Off-peak share", f"{off / (off + pk) * 100 if (off + pk) else 0:.0f}%"),
        ]
        if per100_idx is not None and car in per100_idx.index:
            metrics.append(("Est. distance", f"{int(per100_idx.loc[car, 'Est. km']):,} km"))
            metrics.append(("Cost / 100 km", f"€{per100_idx.loc[car, '€/100km']:.2f}"))
        per_car_ctx.append(
            {
                "name": car,
                "color": CAR_COLOR.get(car, "#2a9d8f"),
                "metrics": metrics,
                "months": [[m, float(c)] for m, c in mbc[mbc["car"] == car][["month", "cost"]].values],
            }
        )

    ov = by_car.assign(
        share=lambda d: d["cost"] / d["cost"].sum() * 100 if d["cost"].sum() else 0
    )
    pdf_ctx = {
        "title": "Car Charging Costs",
        "subtitle": " · ".join(caption) if caption else "EV charging report",
        "kpis": [
            ("Charging sessions", f"{n_sessions}"),
            ("Energy", f"{total_kwh:,.1f} kWh"),
            ("Cost", f"€{total_cost:,.2f}"),
            ("Avg €/session", f"€{avg_cost:,.2f}"),
            ("Effective €/kWh", f"€{eur_per_kwh:,.3f}"),
            ("Off-peak share", f"{off_share:.0f}%"),
        ],
        "overview": [["Car", "Sessions", "kWh", "Cost (€)", "Share"]]
        + [
            [r["car"], int(r["sessions"]), f"{r['energy_kwh']:.1f}", f"{r['cost']:.2f}", f"{r['share']:.0f}%"]
            for _, r in ov.iterrows()
        ],
        "share": [(r["car"], float(r["cost"]), CAR_COLOR.get(r["car"], "#2a9d8f")) for _, r in by_car.iterrows()],
        "tou": [["Period", "kWh", "Price (€/kWh)", "Cost (€)"]]
        + [
            [row["Period"], f"{row['kWh']:.1f}", "" if row["Price (€/kWh)"] is None else f"{row['Price (€/kWh)']:.2f}", f"{row['Cost (€)']:.2f}"]
            for _, row in tou_summary.iterrows()
        ],
        "per_car": per_car_ctx,
    }

    sheets = {
        "Summary": summary,
        "Sessions": show,
        "Monthly": msum,
        "By car": by_car,
        "Daily": daily,
        "Effective price": mprice,
        "Peak vs off-peak": tou_summary,
        "Cost per 100km": per100,
    }
    # Fold in the latest P1 reading (from the Home tab) so the download combines
    # home energy with the charging data.
    r_home = st.session_state.get("p1_reading")
    if r_home is not None:
        sheets["Home energy (P1)"] = p1_details_df(r_home)
        sheets["Home vs car"] = home_vs_car_df(r_home, home_price, total_kwh, total_cost)

    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "⬇️ Excel",
        to_excel(sheets),
        file_name="charging_costs.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    try:
        from report import build_pdf

        d2.download_button(
            "⬇️ PDF report",
            build_pdf(pdf_ctx),
            file_name="charging_report.pdf",
            mime="application/pdf",
        )
    except Exception as exc:  # noqa: BLE001
        d2.warning(f"PDF unavailable: {exc}")
    d3.download_button(
        "⬇️ CSV",
        show.to_csv(index=False).encode("utf-8"),
        file_name="charging_sessions.csv",
        mime="text/csv",
    )

